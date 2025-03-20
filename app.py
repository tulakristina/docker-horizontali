from flask import Flask, render_template_string, request, send_file
from openpyxl import load_workbook
import io

app = Flask(__name__)

# A very simple HTML form for uploading a file
HTML_TEMPLATE = """
<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <title>Horizontali ataskaita</title>
  </head>
  <body>
    <h1>Įkelkite duomenis iš VDA sistemos excel formatu</h1>
    <form action="/" method="post" enctype="multipart/form-data">
      <input type="file" name="source_file" accept=".xlsx" required>
      <br><br>
      <input type="submit" value="Atsisiųsti ataskaitą">
    </form>
  </body>
</html>
"""

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        if "source_file" not in request.files:
            return "Failas neįkeltas", 400

        source_file = request.files["source_file"]

        # Load uploaded workbook using openpyxl (assumes source file has data in D2:D8)
        try:
            source_wb = load_workbook(source_file, data_only=True)
            source_ws = source_wb.active

            # for row in source_ws.iter_rows():
            #   for cell in row:
            #       if cell.value == "":
            #           cell.value = None


            values1 = [source_ws[f"D{row}"].value for row in range(2, 9)]
            values8 = [
                int(source_ws[f"D{row}"].value) if source_ws[f"D{row}"].value else 0
                for row in [10, 11, 9]
            ]
            values10 = [source_ws[f"D{row}"].value for row in [16, 17, 15]] # 10
            values12 = [source_ws[f"D{row}"].value for row in [22, 23, 21]] # 12
            values14 = [source_ws[f"D{row}"].value for row in [28, 29, 27]] # 14
            values16 = [source_ws[f"D{row}"].value for row in [34, 35, 33]] # 16
            values18 = [source_ws[f"D{row}"].value for row in [40, 41, 39]] # 18

            # reiksmes su reziais

            # values11_1 = [source_ws[f"G{row}"].value for row in [19, 20, 18]] 
            # values11_2 = [source_ws[f"H{row}"].value for row in [19, 20, 18]] 
            # values11 = [f"{g}-{h}" for g, h in zip(values11_1, values11_2)]

            # values13_1 = [source_ws[f"G{row}"].value for row in [25, 26, 24]] 
            # values13_2 = [source_ws[f"H{row}"].value for row in [25, 26, 24]] 
            # values13 = [f"{g}-{h}" for g, h in zip(values13_1, values13_2)]

            # values15_1 = [source_ws[f"G{row}"].value for row in [31, 32, 30]] 
            # values15_2 = [source_ws[f"H{row}"].value for row in [31, 32, 30]] 
            # values15 = [f"{g}-{h}" for g, h in zip(values15_1, values15_2)]

            # values17_1 = [source_ws[f"G{row}"].value for row in [37, 38, 36]] 
            # values17_2 = [source_ws[f"H{row}"].value for row in [37, 38, 36]] 
            # values17 = [f"{g}-{h}" for g, h in zip(values17_1, values17_2)]

            # values19_1 = [source_ws[f"G{row}"].value for row in [43, 44, 42]] 
            # values19_2 = [source_ws[f"H{row}"].value for row in [43, 44, 42]] 
            # values19 = [f"{g}-{h}" for g, h in zip(values19_1, values19_2)]

            # values21_1 = [source_ws[f"G{row}"].value for row in [49, 48]] 
            # values21_2 = [source_ws[f"H{row}"].value for row in [49, 48]] 
            # values21 = [f"{g}-{h}" for g, h in zip(values21_1, values21_2)]

            # values22_1 = [source_ws[f"G{row}"].value for row in [51, 50]] 
            # values22_2 = [source_ws[f"H{row}"].value for row in [51, 50]] 
            # values22 = [f"{g}-{h}" for g, h in zip(values22_1, values22_2)]

            # values23_1 = [source_ws[f"G{row}"].value for row in [53, 52]] 
            # values23_2 = [source_ws[f"H{row}"].value for row in [53, 52]] 
            # values23 = [f"{g}-{h}" for g, h in zip(values23_1, values23_2)]

            # values24_1 = [source_ws[f"G{row}"].value for row in [55, 54]] 
            # values24_2 = [source_ws[f"H{row}"].value for row in [55, 54]] 
            # values24 = [f"{g}-{h}" for g, h in zip(values24_1, values24_2)]

            # values25_1 = [source_ws[f"G{row}"].value for row in [57, 56]] 
            # values25_2 = [source_ws[f"H{row}"].value for row in [57, 56]] 
            # values25 = [f"{g}-{h}" for g, h in zip(values25_1, values25_2)]

            # values21_25 = [values21, values22, values23, values24, values25]

            

            # values27_1 = [source_ws[f"G{row}"].value for row in [62, 61]] 
            # values27_2 = [source_ws[f"H{row}"].value for row in [62, 61]] 
            # values27 = [f"{g}-{h}" for g, h in zip(values27_1, values27_2)]

            # values28_1 = [source_ws[f"G{row}"].value for row in [64, 63]] 
            # values28_2 = [source_ws[f"H{row}"].value for row in [64, 63]] 
            # values28 = [f"{g}-{h}" for g, h in zip(values28_1, values28_2)]

            # values29_1 = [source_ws[f"G{row}"].value for row in [66, 65]] 
            # values29_2 = [source_ws[f"H{row}"].value for row in [66, 65]] 
            # values29 = [f"{g}-{h}" for g, h in zip(values29_1, values29_2)]

            # values30_1 = [source_ws[f"G{row}"].value for row in [68, 67]] 
            # values30_2 = [source_ws[f"H{row}"].value for row in [68, 67]] 
            # values30 = [f"{g}-{h}" for g, h in zip(values30_1, values30_2)]

            # values27_30 = [values27, values28, values29, values30]

            def safe_int(value):
                try:
                    return int(value)
                except (TypeError, ValueError):
                    return 0

            values11_1 = [safe_int(source_ws[f"G{row}"].value) for row in [19, 20, 18]]
            values11_2 = [safe_int(source_ws[f"H{row}"].value) for row in [19, 20, 18]]
            values11 = [f"{g}-{h}" for g, h in zip(values11_1, values11_2)]

            values13_1 = [safe_int(source_ws[f"G{row}"].value) for row in [25, 26, 24]]
            values13_2 = [safe_int(source_ws[f"H{row}"].value) for row in [25, 26, 24]]
            values13 = [f"{g}-{h}" for g, h in zip(values13_1, values13_2)]

            values15_1 = [safe_int(source_ws[f"G{row}"].value) for row in [31, 32, 30]]
            values15_2 = [safe_int(source_ws[f"H{row}"].value) for row in [31, 32, 30]]
            values15 = [f"{g}-{h}" for g, h in zip(values15_1, values15_2)]

            values17_1 = [safe_int(source_ws[f"G{row}"].value) for row in [37, 38, 36]]
            values17_2 = [safe_int(source_ws[f"H{row}"].value) for row in [37, 38, 36]]
            values17 = [f"{g}-{h}" for g, h in zip(values17_1, values17_2)]

            values19_1 = [safe_int(source_ws[f"G{row}"].value) for row in [43, 44, 42]]
            values19_2 = [safe_int(source_ws[f"H{row}"].value) for row in [43, 44, 42]]
            values19 = [f"{g}-{h}" for g, h in zip(values19_1, values19_2)]

            values21_1 = [safe_int(source_ws[f"G{row}"].value) for row in [49, 48]]
            values21_2 = [safe_int(source_ws[f"H{row}"].value) for row in [49, 48]]
            values21 = [f"{g}-{h}" for g, h in zip(values21_1, values21_2)]

            values22_1 = [safe_int(source_ws[f"G{row}"].value) for row in [51, 50]]
            values22_2 = [safe_int(source_ws[f"H{row}"].value) for row in [51, 50]]
            values22 = [f"{g}-{h}" for g, h in zip(values22_1, values22_2)]

            values23_1 = [safe_int(source_ws[f"G{row}"].value) for row in [53, 52]]
            values23_2 = [safe_int(source_ws[f"H{row}"].value) for row in [53, 52]]
            values23 = [f"{g}-{h}" for g, h in zip(values23_1, values23_2)]

            values24_1 = [safe_int(source_ws[f"G{row}"].value) for row in [55, 54]]
            values24_2 = [safe_int(source_ws[f"H{row}"].value) for row in [55, 54]]
            values24 = [f"{g}-{h}" for g, h in zip(values24_1, values24_2)]

            values25_1 = [safe_int(source_ws[f"G{row}"].value) for row in [57, 56]]
            values25_2 = [safe_int(source_ws[f"H{row}"].value) for row in [57, 56]]
            values25 = [f"{g}-{h}" for g, h in zip(values25_1, values25_2)]

            values21_25 = [values21, values22, values23, values24, values25]

            values27_1 = [safe_int(source_ws[f"G{row}"].value) for row in [62, 61]]
            values27_2 = [safe_int(source_ws[f"H{row}"].value) for row in [62, 61]]
            values27 = [f"{g}-{h}" for g, h in zip(values27_1, values27_2)]

            values28_1 = [safe_int(source_ws[f"G{row}"].value) for row in [64, 63]]
            values28_2 = [safe_int(source_ws[f"H{row}"].value) for row in [64, 63]]
            values28 = [f"{g}-{h}" for g, h in zip(values28_1, values28_2)]

            values29_1 = [safe_int(source_ws[f"G{row}"].value) for row in [66, 65]]
            values29_2 = [safe_int(source_ws[f"H{row}"].value) for row in [66, 65]]
            values29 = [f"{g}-{h}" for g, h in zip(values29_1, values29_2)]

            values30_1 = [safe_int(source_ws[f"G{row}"].value) for row in [68, 67]]
            values30_2 = [safe_int(source_ws[f"H{row}"].value) for row in [68, 67]]
            values30 = [f"{g}-{h}" for g, h in zip(values30_1, values30_2)]

            values27_30 = [values27, values28, values29, values30]

            # kitas
            values31_43 = [source_ws[f"D{row}"].value for row in range(69, 82)]

            values44_57 = [source_ws[f"D{row}"].value for row in range(82, 96)]

            values58_67 = [source_ws[f"D{row}"].value for row in range(96, 106)]

            values68_74 = [source_ws[f"D{row}"].value if i == 0 else int(source_ws[f"D{row}"].value) 
               for i, row in enumerate(range(106, 113))]

            values75_87 = []
            for row in range(113, 126):
                cell_value = source_ws[f"D{row}"].value
                try:
                    formatted_value = "{:.2f}".format(float(cell_value))
                    values75_87.append(formatted_value)
                except (TypeError, ValueError):
                    values75_87.append("0.00")

            # # # 9
            # # #values9_g_min = min(source_ws[f"G{row}"].value for row in [19, 25, 31, 37, 43]) # gauta
            # # #values9_n_min = min(source_ws[f"G{row}"].value for row in [20, 26, 32, 38, 44]) # nurasyta
            # # #values9_f_min = min(source_ws[f"G{row}"].value for row in [18, 24, 30, 36, 42]) # fondas

            # values = [source_ws[f"G{row}"].value for row in [19, 25, 31, 37, 43] if source_ws[f"G{row}"].value is not None]
            # values9_g_min = min(values) if values else 0  # gauta
            # values = [source_ws[f"G{row}"].value for row in [20, 26, 32, 38, 44] if source_ws[f"G{row}"].value is not None] # fondas
            # values9_n_min = min(values) if values else 0  # nurasyta
            # values = [source_ws[f"G{row}"].value for row in [18, 24, 30, 36, 42] if source_ws[f"G{row}"].value is not None] # fondas
            # values9_f_min = min(values) if values else 0  # fondas

            # # #values9_g_max = max(source_ws[f"H{row}"].value for row in [19, 25, 31, 37, 43]) # gauta
            # # #values9_n_max = max(source_ws[f"H{row}"].value for row in [20, 26, 32, 38, 44]) # nurasyta
            # # #values9_f_max = max(source_ws[f"H{row}"].value for row in [18, 24, 30, 36, 42]) # fondas

            # values = [source_ws[f"H{row}"].value for row in [19, 25, 31, 37, 43] if source_ws[f"H{row}"].value is not None] # gauta
            # values9_g_max = max(values) if values else 0  # gauta
            # values = [source_ws[f"H{row}"].value for row in [20, 26, 32, 38, 44] if source_ws[f"H{row}"].value is not None] # fondas
            # values9_n_max = max(values) if values else 0 # nurasyta
            # values = [source_ws[f"H{row}"].value for row in [18, 24, 30, 36, 42] if source_ws[f"H{row}"].value is not None] # fondas
            # values9_f_max = max(values) if values else 0 # fondas


            
            # values9 = [
            #     f"{values9_g_min}-{values9_g_max}",
            #     f"{values9_n_min}-{values9_n_max}",
            #     f"{values9_f_min}-{values9_f_max}",
            # ]

            # # 20

            # values = [source_ws[f"G{row}"].value for row in [49, 51, 53, 55] if source_ws[f"G{row}"].value is not None]
            # values20_g_min = min(values) if values else 0
            # values = [source_ws[f"G{row}"].value for row in [48, 50, 52, 54] if source_ws[f"G{row}"].value is not None] # fondas
            # values20_f_min = min(values) if values else 0

            # values = [source_ws[f"H{row}"].value for row in [49, 51, 53, 55] if source_ws[f"H{row}"].value is not None] # gauta
            # values20_g_max = max(values) if values else 0
            # values = [source_ws[f"H{row}"].value for row in [48, 50, 52, 54] if source_ws[f"H{row}"].value is not None] # fondas
            # values20_f_max = max(values) if values else 0

            # values20 = [
            #     f"{values20_g_min}-{values20_g_max}",
            #     f"{values20_f_min}-{values20_f_max}",
            # ]

            # # 26

            # values = [source_ws[f"G{row}"].value for row in [62, 64, 66, 68] if source_ws[f"G{row}"].value is not None]
            # values26_g_min = min(values) if values else 0
            # values = [source_ws[f"G{row}"].value for row in [61, 63, 65, 67] if source_ws[f"G{row}"].value is not None] # fondas
            # values26_f_min = min(values) if values else 0

            # values = [source_ws[f"H{row}"].value for row in [62, 64, 66, 68] if source_ws[f"H{row}"].value is not None] # gauta
            # values26_g_max = max(values) if values else 0
            # values = [source_ws[f"H{row}"].value for row in [62, 64, 66, 68] if source_ws[f"H{row}"].value is not None] # fondas
            # values26_f_max = max(values) if values else 0

            # values26 = [
            #     f"{values26_g_min}-{values26_g_max}",
            #     f"{values26_f_min}-{values26_f_max}",
            # ]


        except Exception as e:
            return f":( Nepavyko nuskaityti failo: {e}", 500

        # Load a template workbook (assumes 'tuscias.xlsx' is present in the container)
        try:
            template_wb = load_workbook("tuscias.xlsx")
            template_ws = template_wb.active

            for index, value in enumerate(values1):
                template_ws.cell(row=39, column= 1 + index).value = value # target_row =39 start_col = 1  # Column A is 1

            for index, value in enumerate(values8):
              col = 2 
              row = 48 + index
              for merged_range in template_ws.merged_cells.ranges:
                if template_ws.cell(row=row, column=col).coordinate in merged_range:
                  template_ws.unmerge_cells(str(merged_range))  # Unmerge before writing
                  break

              template_ws.cell(row=row, column= col).value = value # target_row =48 start_col = 2  # Column A is 1
              #if index == 0:  # Example: Re-merge for first value
              template_ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)

            for index, value in enumerate(values10):
              col = 6 
              row = 48 + index
              template_ws.cell(row=row, column= col).value = value 

            for index, value in enumerate(values12):
              col = 8 
              row = 48 + index
              template_ws.cell(row=row, column= col).value = value 
            
            for index, value in enumerate(values14):
              col = 10 
              row = 48 + index
              template_ws.cell(row=row, column= col).value = value

            for index, value in enumerate(values16):
              col = 12 
              row = 48 + index
              template_ws.cell(row=row, column= col).value = value

            for index, value in enumerate(values18):
              col = 14 
              row = 48 + index
              template_ws.cell(row=row, column= col).value = value

          # nelyginiai reziai

            for index, value in enumerate(values11):
              col = 7 
              row = 48 + index
              template_ws.cell(row=row, column= col).value = value

            for index, value in enumerate(values13):
              col = 9 
              row = 48 + index
              template_ws.cell(row=row, column= col).value = value

            for index, value in enumerate(values15):
              col = 11 
              row = 48 + index
              template_ws.cell(row=row, column= col).value = value

            for index, value in enumerate(values17):
              col = 13 
              row = 48 + index
              template_ws.cell(row=row, column= col).value = value

            for index, value in enumerate(values19):
              col = 15 
              row = 48 + index
              template_ws.cell(row=row, column= col).value = value

            # # 9 reziai, per kelis langelius

            # for index, value in enumerate(values9):
            #   col = 4 
            #   row = 48 + index

            #   for merged_range in template_ws.merged_cells.ranges:
            #     if template_ws.cell(row=row, column=col).coordinate in merged_range:
            #       template_ws.unmerge_cells(str(merged_range))  # Unmerge before writing
            #       break

            #   template_ws.cell(row=row, column= col).value = value
            #   template_ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)

            # # 20 reziai per kelis langelius 

            # rows = [58, 59, 60]

            # for index in range(3):
            #   col = 2 
            #   row = 58 + index
            #   for merged_range in template_ws.merged_cells.ranges:
            #     if template_ws.cell(row=row, column=col).coordinate in merged_range:
            #       template_ws.unmerge_cells(str(merged_range))  # Unmerge before writing
            #       break

            # template_ws.cell(row=58, column= 2).value = values20[0]
            # template_ws.cell(row=59, column= 2).value = 0
            # template_ws.cell(row=60, column= 2).value = values20[1]

            # for index in range(3):
            #   col = 2 
            #   row = 58 + index
            #   template_ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)

            # # 26

            
            # template_ws.cell(row=58, column= 10).value = values26[0]
            # template_ws.cell(row=59, column= 10).value = 0
            # template_ws.cell(row=60, column= 10).value = values26[1]  

            # El. dokumentų fondas

            for index in range(5):
              col = 5 + index
              template_ws.cell(row=58, column= col).value = values21_25[index][0]
              template_ws.cell(row=59, column= col).value = 0
              template_ws.cell(row=60, column= col).value = values21_25[index][1]

            for index in range(4):
              col = 11 + index
              template_ws.cell(row=58, column= col).value = values27_30[index][0]
              template_ws.cell(row=59, column= col).value = 0
              template_ws.cell(row=60, column= col).value = values27_30[index][1]  

            # istekliai ir paslaugos

            for index, value in enumerate(values31_43):
              template_ws.cell(row=69, column= 1 + index).value = value

            for index, value in enumerate(values44_57):
              template_ws.cell(row=79, column= 1 + index).value = value  
              
            col_index = 1

            for value in values58_67:
              row = 87
              merged = False
              for merged_range in template_ws.merged_cells.ranges:
                if template_ws.cell(row=row, column=col_index).coordinate in merged_range:
                  template_ws.unmerge_cells(str(merged_range))  # Unmerge before writing
                  merged = True
                  break
              template_ws.cell(row=row, column= col_index).value = value  
              if merged:
                template_ws.merge_cells(start_row=row, start_column=col_index, end_row=row, end_column=col_index+1)
              col_index += 2 if merged else 1

            # darbuotojai
            col_index = 1
            for value in values68_74:
              row= 96
              merged = False
              for merged_range in template_ws.merged_cells.ranges:
                if template_ws.cell(row=row, column=col_index).coordinate in merged_range:
                  template_ws.unmerge_cells(str(merged_range))  # Unmerge before writing
                  merged=True
                  break
              template_ws.cell(row=row, column= col_index).value = value
              if merged:
                template_ws.merge_cells(start_row=row, start_column=col_index, end_row=row, end_column=col_index+1)
              col_index += 2 if merged else 1

              #  finansavimas ir išlaidos
            for index, value in enumerate(values75_87):
              template_ws.cell(row=107, column= 1 + index).value = value


        except Exception as e:
            return f"Error processing template: {e}", 500

        # Save modified workbook into a BytesIO stream and send it for download
        output = io.BytesIO()
        template_wb.save(output)
        output.seek(0)
        return send_file(output,
                         as_attachment=True,
                         download_name="ataskaita.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return render_template_string(HTML_TEMPLATE)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080)
